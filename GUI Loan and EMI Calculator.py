# Python GUI program to calculate loan and EMI

import tkinter as tk
from tkinter import ttk  # To create GUI widgets
import openpyxl as op  # To input data into the Excel file

# function which is called by the calculate button


def emi():
    loan = loan_type.get()
    p = float(principal.get())
    t = float(tenure.get())
    r = float(interest.get())
    i = (r / 12) / 100

    m = (1 + i) ** t
    installment = (p * i) * (m / (m - 1))
    amt = installment * t

    # average interest rates are enterd for each type of loan

    if loan == "Personal Loan":
        if r <= 13:
            banner.configure(
                text="EMI : ₹ "
                + str(round(installment, 4))
                + "\n"
                + " Total Amount = ₹"
                + str(round(amt, 4)),
                bg="green",
                font="bold 35",
            )
        elif r >= 14 and r <= 19:
            banner.configure(
                text="EMI : ₹ "
                + str(round(installment, 4))
                + "\n"
                + " Total Amount = ₹"
                + str(round(amt, 4)),
                bg="orange",
                font="bold 35",
            )
        else:
            banner.configure(
                text="EMI : ₹ "
                + str(round(installment, 4))
                + "\n"
                + " Total Amount = ₹"
                + str(round(amt, 4)),
                bg="red",
                font="bold 35",
            )
    elif loan == "Vehicle Loan":
        if r <= 9:
            banner.configure(
                text="EMI : ₹ "
                + str(round(installment, 4))
                + "\n"
                + " Total Amount = ₹"
                + str(round(amt, 4)),
                bg="green",
                font="bold 35",
            )
        elif r >= 10 and r <= 11:
            banner.configure(
                text="EMI : ₹ "
                + str(round(installment, 4))
                + "\n"
                + " Total Amount = ₹"
                + str(round(amt, 4)),
                bg="orange",
                font="bold 35",
            )
        else:
            banner.configure(
                text="EMI : ₹ "
                + str(round(installment, 4))
                + "\n"
                + " Total Amount = ₹"
                + str(round(amt, 4)),
                bg="red",
                font="bold 35",
            )
    else:
        if r <= 7:
            banner.configure(
                text="EMI : ₹ "
                + str(round(installment, 4))
                + "\n"
                + " Total Amount = ₹"
                + str(round(amt, 4)),
                bg="green",
                font="bold 35",
            )
        elif r >= 8 and r <= 9:
            banner.configure(
                text="EMI : ₹ "
                + str(round(installment, 4))
                + "\n"
                + " Total Amount = ₹"
                + str(round(amt, 4)),
                bg="orange",
                font="bold 35",
            )
        else:
            banner.configure(
                text="EMI : ₹ "
                + str(round(installment, 4))
                + "\n"
                + " Total Amount = ₹"
                + str(round(amt, 4)),
                bg="red",
                font="bold 35",
            )
    wb = op.Workbook()
    ws = wb.active
    ws["A1"] = "MONTH"
    ws["B1"] = "EMI"
    ws["C1"] = "OUTSTANDING"

    # code which calculte the outstanding balance amount for each month which i used
    # inputing into the excel file
    for month in range(1, int(t) + 1):
        outstanding_amt = amt - (installment * month)
        new_data = [month, installment, outstanding_amt]
        ws.append(new_data)
    wb.save("C:\\Users\\hp\\Desktop\\Python LA2\\demonstration.xlsx")


window = tk.Tk()  # creates object Tk
window.title("EMI Calculator")


main_frame = tk.Frame(window)
main_frame.pack()

# Frames are created for dividing the program into to sections
frame_1 = tk.LabelFrame(main_frame)
frame_1.grid(row=0, column=0, padx=80, pady=20)

frame_2 = tk.LabelFrame(main_frame)
frame_2.grid(row=1, column=0)

frame_3 = tk.LabelFrame(main_frame)
frame_3.grid(row=2, column=0)


banner = tk.Label(
    frame_1, text="Easy EMI", bg="blue", fg="black", padx=100, pady=10, font="bold 70"
)
banner.grid(row=0, column=0, padx=100, pady=20, sticky="news")

# Buttons , entry box are placed into the window
text_1_label = tk.Label(frame_2, text="Select the type of Loan: ", font="bold 20")
text_1_label.grid(row=0, column=0, padx=40, pady=20)
loan_type = ttk.Combobox(
    frame_2, values=["Personal Loan", "Vehicle Loan", "Home Loan"], font="30"
)
loan_type.grid(row=0, column=1, padx=40, pady=20)

text_2_label = tk.Label(frame_2, text="Enter the Principal Amount: ", font="bold 20")
text_2_label.grid(row=1, column=0, padx=40, pady=20)
principal = tk.Entry(frame_2, font="30")
principal.grid(row=1, column=1, padx=40, pady=20)

text_3_label = tk.Label(frame_2, text="Enter the Interest Rate: ", font="bold 20")
text_3_label.grid(row=2, column=0, padx=40, pady=20)
interest = tk.Entry(frame_2, font="30")
interest.grid(row=2, column=1, padx=40, pady=20)

text_4_label = tk.Label(frame_2, text="Enter the Tenure in Months: ", font="bold 20")
text_4_label.grid(row=3, column=0, padx=40, pady=20)
tenure = tk.Entry(frame_2, font="30")
tenure.grid(row=3, column=1, padx=40, pady=20)

calculate_button = tk.Button(
    main_frame, text="Calculate", bg="black", fg="white", font="bold 20", command=emi
)
calculate_button.grid(row=4, column=0, padx=40, pady=20)

window.mainloop()
